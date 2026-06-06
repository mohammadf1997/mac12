import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import matplotlib.pyplot as plt
from tkcalendar import DateEntry

# -------- پیدا کردن مسیر فایل برنامه --------
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_data_file_path():
    """Get path to data file (stored next to the executable)"""
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        application_path = os.path.dirname(sys.executable)
    else:
        # Running as script
        application_path = os.path.dirname(os.path.abspath(__file__))
    
    # Create data folder if it doesn't exist
    data_folder = os.path.join(application_path, "InternshipData")
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)
    
    return os.path.join(data_folder, "internship_tracker.xlsx")

# -------- Settings --------
TOTAL_REQUIRED_HOURS = 450
TOTAL_REQUIRED_MINUTES = TOTAL_REQUIRED_HOURS * 60

# فایل الان در پوشه برنامه ذخیره میشه
FILE = get_data_file_path()

# -------- Helper Functions --------
def minutes_to_hms(total_minutes):
    hours = int(total_minutes // 60)
    minutes = int(total_minutes % 60)
    return hours, minutes

def format_time_display(total_minutes):
    hours, minutes = minutes_to_hms(total_minutes)
    return f"{hours:02d}:{minutes:02d}"

def format_hours_minutes(hours, minutes):
    h = int(hours) if hours else 0
    m = int(minutes) if minutes else 0
    return f"{h:02d}:{m:02d}"

def calculate_hours_minutes(start, end):
    fmt = "%H:%M"
    s = datetime.strptime(start, fmt)
    e = datetime.strptime(end, fmt)
    
    if e < s:
        raise ValueError("End time must be after start time")
    
    diff_seconds = (e - s).seconds
    hours = diff_seconds // 3600
    minutes = (diff_seconds % 3600) // 60
    
    return hours, minutes

# -------- Initialize Excel File --------
def init_file():
    # اطمینان از وجود پوشه داده
    data_dir = os.path.dirname(FILE)
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
    
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Internship Records"
        headers = ["Date", "Title", "Start Time", "End Time", "Hours", "Minutes"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        wb.save(FILE)
        return True
    return False

def save_data(date, title, start, end, hours, minutes):
    try:
        # اطمینان از وجود فایل
        init_file()
        
        wb = load_workbook(FILE)
        ws = wb.active
        ws.append([date, title, start, end, int(hours), int(minutes)])
        wb.save(FILE)
        return True
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data: {str(e)}")
        return False

def load_data():
    try:
        if not os.path.exists(FILE):
            init_file()
            return []
        
        wb = load_workbook(FILE)
        ws = wb.active
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                if len(row) >= 6:
                    data.append(list(row[:6]))
                else:
                    data.append([row[0], row[1], row[2], row[3], row[4] if row[4] else 0, 0])
        
        return data
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load data: {str(e)}")
        return []

def get_total_minutes(data):
    total_minutes = 0
    for record in data:
        if len(record) >= 6:
            hours = int(record[4]) if record[4] else 0
            minutes = int(record[5]) if record[5] else 0
            total_minutes += (hours * 60) + minutes
        else:
            total_minutes += int(record[4] * 60)
    return total_minutes

# -------- Delete Record --------
def delete_record():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Warning", "Please select a record to delete")
        return
    
    if messagebox.askyesno("Confirm", "Are you sure you want to delete this record?"):
        try:
            wb = load_workbook(FILE)
            ws = wb.active
            
            for item in selected:
                values = tree.item(item, 'values')
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if row[0] == values[0] and row[1] == values[1] and row[2] == values[2]:
                        ws.delete_rows(row_idx)
                        break
            
            wb.save(FILE)
            update_ui()
            messagebox.showinfo("Success", "Record deleted successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete: {str(e)}")

# -------- Edit Record --------
def edit_record():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Warning", "Please select a record to edit")
        return
    
    values = tree.item(selected[0], 'values')
    
    edit_win = tk.Toplevel(root)
    edit_win.title("Edit Record")
    edit_win.geometry("500x450")
    edit_win.configure(bg="#1e1e2f")
    edit_win.transient(root)
    edit_win.grab_set()
    
    tk.Label(edit_win, text="Date:", fg="white", bg="#1e1e2f", font=("Arial", 10, "bold")).pack(pady=5)
    date_entry = DateEntry(edit_win, width=20, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
    date_entry.set_date(values[0])
    date_entry.pack(pady=5)
    
    tk.Label(edit_win, text="Title:", fg="white", bg="#1e1e2f", font=("Arial", 10, "bold")).pack(pady=5)
    title_entry = tk.Entry(edit_win, width=40, font=("Arial", 10))
    title_entry.insert(0, values[1])
    title_entry.pack(pady=5)
    
    time_frame = tk.Frame(edit_win, bg="#1e1e2f")
    time_frame.pack(pady=10)
    
    tk.Label(time_frame, text="Start Time:", fg="white", bg="#1e1e2f", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=5)
    start_hour = tk.Spinbox(time_frame, from_=0, to=23, width=5, format="%02.0f", font=("Arial", 12))
    start_hour.delete(0, tk.END)
    start_hour.insert(0, values[2].split(":")[0] if ":" in values[2] else "09")
    start_hour.grid(row=0, column=1, padx=2)
    tk.Label(time_frame, text=":", fg="white", bg="#1e1e2f", font=("Arial", 14, "bold")).grid(row=0, column=2)
    start_minute = tk.Spinbox(time_frame, from_=0, to=59, width=5, format="%02.0f", font=("Arial", 12))
    start_minute.delete(0, tk.END)
    start_minute.insert(0, values[2].split(":")[1] if ":" in values[2] else "00")
    start_minute.grid(row=0, column=3, padx=2)
    
    tk.Label(time_frame, text="    End Time:", fg="white", bg="#1e1e2f", font=("Arial", 10, "bold")).grid(row=0, column=4, padx=5)
    end_hour = tk.Spinbox(time_frame, from_=0, to=23, width=5, format="%02.0f", font=("Arial", 12))
    end_hour.delete(0, tk.END)
    end_hour.insert(0, values[3].split(":")[0] if ":" in values[3] else "17")
    end_hour.grid(row=0, column=5, padx=2)
    tk.Label(time_frame, text=":", fg="white", bg="#1e1e2f", font=("Arial", 14, "bold")).grid(row=0, column=6)
    end_minute = tk.Spinbox(time_frame, from_=0, to=59, width=5, format="%02.0f", font=("Arial", 12))
    end_minute.delete(0, tk.END)
    end_minute.insert(0, values[3].split(":")[1] if ":" in values[3] else "00")
    end_minute.grid(row=0, column=7, padx=2)
    
    def save_edit():
        try:
            start_time = f"{int(start_hour.get()):02d}:{int(start_minute.get()):02d}"
            end_time = f"{int(end_hour.get()):02d}:{int(end_minute.get()):02d}"
            hours, minutes = calculate_hours_minutes(start_time, end_time)
            
            wb = load_workbook(FILE)
            ws = wb.active
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[0] == values[0] and row[1] == values[1] and row[2] == values[2]:
                    ws.delete_rows(row_idx)
                    break
            
            ws.append([date_entry.get(), title_entry.get(), start_time, end_time, int(hours), int(minutes)])
            wb.save(FILE)
            
            edit_win.destroy()
            update_ui()
            messagebox.showinfo("Success", "Record edited successfully")
        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    tk.Button(edit_win, text="Save Changes", bg="#4CAF50", fg="white", command=save_edit, width=20, height=2, font=("Arial", 10, "bold")).pack(pady=20)

# -------- Add Record --------
def add_record():
    try:
        date = e_date.get()
        title = e_title.get().strip()
        start_hour = int(spin_start_hour.get())
        start_minute = int(spin_start_minute.get())
        end_hour = int(spin_end_hour.get())
        end_minute = int(spin_end_minute.get())
        
        start_time = f"{start_hour:02d}:{start_minute:02d}"
        end_time = f"{end_hour:02d}:{end_minute:02d}"

        if not date or not title:
            messagebox.showerror("Error", "Please fill all fields")
            return

        hours, minutes = calculate_hours_minutes(start_time, end_time)
        
        if hours == 0 and minutes == 0:
            messagebox.showerror("Error", "Work time cannot be zero")
            return

        if save_data(date, title, start_time, end_time, hours, minutes):
            update_ui()
            clear_inputs()
            messagebox.showinfo("Success", f"Record saved: {format_hours_minutes(hours, minutes)}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# -------- Update UI --------
def update_ui():
    for item in tree.get_children():
        tree.delete(item)

    data = load_data()
    total_minutes = get_total_minutes(data)
    remain_minutes = TOTAL_REQUIRED_MINUTES - total_minutes
    percent = (total_minutes / TOTAL_REQUIRED_MINUTES) * 100 if TOTAL_REQUIRED_MINUTES > 0 else 0

    for record in data:
        if len(record) >= 6:
            h = int(record[4]) if record[4] else 0
            m = int(record[5]) if record[5] else 0
            display_time = format_hours_minutes(h, m)
        else:
            h, m = minutes_to_hms(int(record[4] * 60))
            display_time = format_hours_minutes(h, m)
        
        tree.insert("", "end", values=[record[0], record[1], record[2], record[3], display_time])

    total_display = format_time_display(total_minutes)
    lbl_total.config(text=f"Total Hours: {total_display} / {TOTAL_REQUIRED_HOURS:02d}:00")
    
    if remain_minutes <= 0:
        overtime_minutes = abs(remain_minutes)
        overtime_display = format_time_display(overtime_minutes)
        lbl_remain.config(text=f"✅ Completed! Overtime: {overtime_display}")
        lbl_remain.config(fg="#4CAF50")
    else:
        remain_display = format_time_display(remain_minutes)
        lbl_remain.config(text=f"Remaining: {remain_display}")
        lbl_remain.config(fg="#FF5252")
    
    progress_bar['value'] = min(percent, 100)
    lbl_percent.config(text=f"{min(percent, 100):.1f}%")
    
    if percent >= 100:
        progress_bar['style'] = "green.Horizontal.TProgressbar"
    elif percent >= 70:
        progress_bar['style'] = "blue.Horizontal.TProgressbar"
    elif percent >= 40:
        progress_bar['style'] = "orange.Horizontal.TProgressbar"
    else:
        progress_bar['style'] = "red.Horizontal.TProgressbar"

def clear_inputs():
    e_date.set_date(datetime.now())
    e_title.delete(0, tk.END)
    spin_start_hour.delete(0, tk.END)
    spin_start_hour.insert(0, "09")
    spin_start_minute.delete(0, tk.END)
    spin_start_minute.insert(0, "00")
    spin_end_hour.delete(0, tk.END)
    spin_end_hour.insert(0, "17")
    spin_end_minute.delete(0, tk.END)
    spin_end_minute.insert(0, "00")

def create_time_picker(parent, default_hour=9, default_minute=0):
    frame = tk.Frame(parent, bg="#2d2d3f")
    
    hour_spin = tk.Spinbox(frame, from_=0, to=23, width=5, format="%02.0f", font=("Arial", 14, "bold"), justify="center")
    hour_spin.delete(0, tk.END)
    hour_spin.insert(0, f"{default_hour:02d}")
    hour_spin.pack(side=tk.LEFT, padx=2)
    
    tk.Label(frame, text=":", fg="white", bg="#2d2d3f", font=("Arial", 16, "bold")).pack(side=tk.LEFT)
    
    minute_spin = tk.Spinbox(frame, from_=0, to=59, width=5, format="%02.0f", font=("Arial", 14, "bold"), justify="center")
    minute_spin.delete(0, tk.END)
    minute_spin.insert(0, f"{default_minute:02d}")
    minute_spin.pack(side=tk.LEFT, padx=2)
    
    return frame, hour_spin, minute_spin

# -------- Chart Functions --------
def show_chart():
    data = load_data()
    if not data:
        messagebox.showwarning("Warning", "No data available to display")
        return
    
    total_minutes = get_total_minutes(data)
    percent = (total_minutes / TOTAL_REQUIRED_MINUTES) * 100 if TOTAL_REQUIRED_MINUTES > 0 else 0
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
    
    remaining_minutes = max(0, TOTAL_REQUIRED_MINUTES - total_minutes)
    if total_minutes >= TOTAL_REQUIRED_MINUTES:
        overtime_minutes = total_minutes - TOTAL_REQUIRED_MINUTES
        sizes = [TOTAL_REQUIRED_MINUTES, overtime_minutes]
        labels = [f'Target Achieved\n({TOTAL_REQUIRED_HOURS:02d}:00)', f'Overtime\n({format_time_display(overtime_minutes)})']
        colors = ['#4CAF50', '#FF9800']
    else:
        sizes = [total_minutes, remaining_minutes]
        labels = [f'Worked\n({format_time_display(total_minutes)})', f'Remaining\n({format_time_display(remaining_minutes)})']
        colors = ['#4CAF50', '#FF5252']
    
    ax1.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    ax1.set_title(f'Progress: {percent:.1f}%', fontsize=12, fontweight='bold')
    
    categories = ['Hours Worked', f'Target\n({TOTAL_REQUIRED_HOURS:02d}:00)']
    values = [total_minutes / 60, TOTAL_REQUIRED_HOURS]
    bars = ax2.bar(categories, values, color=['#2196F3', '#FF9800'], edgecolor='black', linewidth=1.5)
    ax2.set_ylabel('Hours', fontsize=11, fontweight='bold')
    ax2.set_title('Work Hours Comparison', fontsize=12, fontweight='bold')
    ax2.grid(axis='y', alpha=0.3)
    
    for bar, value in zip(bars, values):
        height = bar.get_height()
        if value == total_minutes / 60:
            label = format_time_display(total_minutes)
        else:
            label = f"{TOTAL_REQUIRED_HOURS:02d}:00"
        ax2.text(bar.get_x() + bar.get_width()/2., height, label, ha='center', va='bottom', fontsize=11, fontweight='bold')
    
    plt.suptitle(f'Internship Progress Report\nTotal: {format_time_display(total_minutes)}', fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.show()

# -------- Export to Excel --------
def export_to_excel():
    data = load_data()
    if not data:
        messagebox.showwarning("Warning", "No data available to export")
        return
    
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"internship_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    
    if not file_path:
        return
    
    try:
        wb = Workbook()
        
        summary_ws = wb.active
        summary_ws.title = "Summary"
        
        total_minutes = get_total_minutes(data)
        remaining_minutes = max(0, TOTAL_REQUIRED_MINUTES - total_minutes)
        overtime_minutes = max(0, total_minutes - TOTAL_REQUIRED_MINUTES)
        percent = (total_minutes / TOTAL_REQUIRED_MINUTES) * 100 if TOTAL_REQUIRED_MINUTES > 0 else 0
        
        summary_ws['A1'] = "INTERNSHIP PROGRESS REPORT"
        summary_ws['A1'].font = Font(size=18, bold=True, color="1B5E20")
        summary_ws.merge_cells('A1:D1')
        summary_ws['A1'].alignment = Alignment(horizontal="center")
        
        summary_ws['A3'] = "Report Generated:"
        summary_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary_ws['A3'].font = Font(bold=True)
        
        metrics = [
            ["Internship Target", f"{TOTAL_REQUIRED_HOURS:02d}:00"],
            ["Total Hours Worked", format_time_display(total_minutes)],
            ["Remaining Hours", format_time_display(remaining_minutes) if remaining_minutes > 0 else "00:00"],
            ["Overtime Hours", format_time_display(overtime_minutes) if overtime_minutes > 0 else "00:00"],
            ["Completion Percentage", f"{percent:.1f}%"],
            ["Status", "COMPLETED" if total_minutes >= TOTAL_REQUIRED_MINUTES else "IN PROGRESS"]
        ]
        
        for idx, (label, value) in enumerate(metrics, start=5):
            summary_ws[f'A{idx}'] = label
            summary_ws[f'B{idx}'] = value
            summary_ws[f'A{idx}'].font = Font(bold=True, size=11)
            summary_ws[f'A{idx}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 20
        
        details_ws = wb.create_sheet("All Records")
        headers = ["#", "Date", "Title", "Start", "End", "Hours Worked"]
        details_ws.append(headers)
        
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col, header in enumerate(headers, 1):
            cell = details_ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        for idx, record in enumerate(sorted(data, key=lambda x: x[0]), 1):
            if len(record) >= 6:
                h = int(record[4]) if record[4] else 0
                m = int(record[5]) if record[5] else 0
                time_str = format_hours_minutes(h, m)
            else:
                h, m = minutes_to_hms(int(record[4] * 60))
                time_str = format_hours_minutes(h, m)
            
            details_ws.append([idx, record[0], record[1], record[2], record[3], time_str])
            
            for col in range(1, 7):
                cell = details_ws.cell(row=idx + 1, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
        
        details_ws.column_dimensions['A'].width = 6
        details_ws.column_dimensions['B'].width = 15
        details_ws.column_dimensions['C'].width = 35
        details_ws.column_dimensions['D'].width = 12
        details_ws.column_dimensions['E'].width = 12
        details_ws.column_dimensions['F'].width = 15
        
        wb.save(file_path)
        messagebox.showinfo("Success", f"Excel report exported successfully!\n\nFile saved at:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export: {str(e)}")

# -------- Main UI --------
root = None
e_date = None
e_title = None
spin_start_hour = None
spin_start_minute = None
spin_end_hour = None
spin_end_minute = None
tree = None
lbl_total = None
lbl_remain = None
progress_bar = None
lbl_percent = None

def main():
    global root, e_date, e_title, spin_start_hour, spin_start_minute, spin_end_hour, spin_end_minute
    global tree, lbl_total, lbl_remain, progress_bar, lbl_percent
    
    # Initialize file and data folder
    init_file()
    
    root = tk.Tk()
    root.title("🔥 Professional Internship Tracker")
    root.geometry("950x750")
    root.configure(bg="#1e1e2f")
    
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("red.Horizontal.TProgressbar", background='#FF5252', troughcolor='#333333')
    style.configure("orange.Horizontal.TProgressbar", background='#FF9800', troughcolor='#333333')
    style.configure("blue.Horizontal.TProgressbar", background='#2196F3', troughcolor='#333333')
    style.configure("green.Horizontal.TProgressbar", background='#4CAF50', troughcolor='#333333')
    
    frame = tk.LabelFrame(root, text="Add New Record", bg="#2d2d3f", fg="white", font=("Arial", 12, "bold"), padx=20, pady=15)
    frame.pack(pady=10, padx=20, fill="x")
    
    tk.Label(frame, text="Date:", fg="white", bg="#2d2d3f", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5, pady=5)
    e_date = DateEntry(frame, width=20, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd', font=("Arial", 10))
    e_date.grid(row=0, column=1, padx=5, pady=5, sticky="w")
    
    tk.Label(frame, text="Title / Description:", fg="white", bg="#2d2d3f", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky="e", padx=5, pady=5)
    e_title = tk.Entry(frame, width=50, font=("Arial", 10))
    e_title.grid(row=1, column=1, padx=5, pady=5, sticky="w")
    
    tk.Label(frame, text="Start Time:", fg="white", bg="#2d2d3f", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="e", padx=5, pady=5)
    start_frame, spin_start_hour, spin_start_minute = create_time_picker(frame, 9, 0)
    start_frame.grid(row=2, column=1, padx=5, pady=5, sticky="w")
    
    tk.Label(frame, text="End Time:", fg="white", bg="#2d2d3f", font=("Arial", 10, "bold")).grid(row=3, column=0, sticky="e", padx=5, pady=5)
    end_frame, spin_end_hour, spin_end_minute = create_time_picker(frame, 17, 0)
    end_frame.grid(row=3, column=1, padx=5, pady=5, sticky="w")
    
    button_frame = tk.Frame(root, bg="#1e1e2f")
    button_frame.pack(pady=10)
    
    tk.Button(button_frame, text="➕ Add Record", bg="#4CAF50", fg="white", command=add_record, width=16, height=2, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="✏️ Edit Record", bg="#FF9800", fg="white", command=edit_record, width=16, height=2, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="❌ Delete Record", bg="#F44336", fg="white", command=delete_record, width=16, height=2, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="📊 Show Chart", bg="#9C27B0", fg="white", command=show_chart, width=16, height=2, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="📁 Export to Excel", bg="#2196F3", fg="white", command=export_to_excel, width=16, height=2, font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
    
    progress_frame = tk.Frame(root, bg="#1e1e2f")
    progress_frame.pack(pady=10, padx=20, fill="x")
    
    tk.Label(progress_frame, text="Overall Progress:", fg="white", bg="#1e1e2f", font=("Arial", 11, "bold")).pack()
    progress_bar = ttk.Progressbar(progress_frame, length=600, mode='determinate', style="red.Horizontal.TProgressbar")
    progress_bar.pack(pady=5)
    lbl_percent = tk.Label(progress_frame, text="0%", fg="white", bg="#1e1e2f", font=("Arial", 12, "bold"))
    lbl_percent.pack()
    
    tree_frame = tk.Frame(root, bg="#1e1e2f")
    tree_frame.pack(pady=10, padx=20, fill="both", expand=True)
    
    cols = ("Date", "Title", "Start", "End", "Hours")
    tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=12)
    
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=180 if col == "Title" else 120)
    
    vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side=tk.LEFT, fill="both", expand=True)
    vsb.pack(side=tk.RIGHT, fill="y")
    
    stats_frame = tk.Frame(root, bg="#1e1e2f")
    stats_frame.pack(pady=10)
    
    lbl_total = tk.Label(stats_frame, text=f"Total Hours: 00:00 / {TOTAL_REQUIRED_HOURS:02d}:00", fg="#4CAF50", bg="#1e1e2f", font=("Arial", 14, "bold"))
    lbl_total.pack()
    
    lbl_remain = tk.Label(stats_frame, text=f"Remaining: {TOTAL_REQUIRED_HOURS:02d}:00", fg="#FF5252", bg="#1e1e2f", font=("Arial", 12))
    lbl_remain.pack()
    
    update_ui()
    root.mainloop()

if __name__ == "__main__":
    main()